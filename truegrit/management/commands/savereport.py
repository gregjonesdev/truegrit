import os
import csv
import platform
import requests
import subprocess
from openpyxl import load_workbook
from ipaddress import ip_address
from django.core.exceptions import ObjectDoesNotExist

from requests.auth import HTTPDigestAuth
from django.core.management.base import BaseCommand

from truegrit.models import Network, Camera


username = 'root'
password = 'h3bc4m3r4'

GREEN = '\033[92m'
RESET = '\033[0m'

# Define file paths
csv_file_path = r"C:\Users\gregoryjones\OneDrive - Preferred Technologies, LLC\Documents\My Data Sources\Devices.csv"
excel_file_path = r"C:\Users\gregoryjones\OneDrive - Preferred Technologies, LLC\Book3.xlsx"

# List to store tuples
class Command(BaseCommand):  

    def ip_ranges(self, ip_list):
        # Sort IPs numerically
        ip_list = sorted(ip_list, key=lambda ip: int(ip_address(ip)))
        ranges = ""
        start = ip_list[0]
        end = start

        for i in range(1, len(ip_list)):
            
            current_ip = ip_list[i]
            previous_ip = ip_list[i - 1]
            
            # Check if the current IP is consecutive with the previous IP
            if int(ip_address(current_ip)) == int(ip_address(previous_ip)) + 1:
                end = current_ip  # Update the end of the current range
            else:
                # Append range to the string without initial comma
                if ranges:  # Add a comma only if ranges is not empty
                    ranges += ","
                    
                if start == end:
                    ranges += f"{start}"
                else:
                    formatted_end = end.split(".")[-1]
                    ranges += f"{start}-{formatted_end}"
                    
                # Start a new range
                start = current_ip
                end = current_ip

        # Append the final range
        if ranges:
            ranges += ","
        if start == end:
            ranges += f"{start}"
        else:
            formatted_end = end.split(".")[-1]
            ranges += f"{start}-{formatted_end}"

        return ranges

    def print_unassigned_ip_addresses(self, bu_identifier):
        unassigned_cameras = {}
        dhcp_count = {}
        for camera in Camera.objects.filter(
            network__business_unit__identifier=bu_identifier,
            mac_address__isnull=True):
            if not camera.model.name in unassigned_cameras.keys():
                unassigned_cameras[camera.model.name] = []
                dhcp_count[camera.model.name] = 0
            if camera.ip_address:
                unassigned_cameras[camera.model.name].append(camera.ip_address)
            else:
                dhcp_count[camera.model.name] += 1    

        print("\nAvailable IP addresses:\n")

        for each in unassigned_cameras.keys():
            ip_ranges = self.ip_ranges(unassigned_cameras[each])
            print(f"{each}:\n")
            if dhcp_count[each] > 0: 
                print(f"\tDHCP: {dhcp_count[each]} Camera(s)")
            print(f"\tStatic: {len(unassigned_cameras[each])} Camera(s)\n")
            print(f"\t{ip_ranges}\n")
            
        print("\nNext steps:")   
        print("1. Unbox cameras and connect to switch.")
        print("2. Add cameras into Axis Device Manager.")
        print("2. Update firmware and assign IP addresses when required.")
        input("3. Generate device report and press <Enter> to continue.\n\n")

    def can_ping(self, ip_address):
        response = subprocess.run(["ping", ip_address], capture_output=True, text=True)
        # Check the output
        return "Reply from" in response.stdout

    def clear_screen(self):
        # move this to common file
        if platform.system() == "Windows":
            os.system("cls")
        else:
            os.system("clear")

    def get_network_from_bu(self, bu_identifier):
        return Network.objects.get(business_unit__identifier=bu_identifier)


    def get_attribute_string(self, ip_address, value_name):
        # works
        url = "http://{}/axis-cgi/param.cgi?action=list&group={}".format(ip_address, value_name)
        try:    
            response = requests.get(url, auth=HTTPDigestAuth(username, password))
            return response.content
        except requests.exceptions.ConnectionError:
            print("\nUnable to connect to {}".format(ip_address))

    def get_attribute_from_ip(self, ip_address, attribute_name):
        attribute_string = str(self.get_attribute_string(
            ip_address,
            attribute_name))
        return self.extract_value(attribute_string)

    def extract_value(self, input_string):
        parts = input_string.split('=') 
        if len(parts) > 1:
            value = parts[1] 
            return value.split("\\")[0].strip().replace(":", "")    

    def is_dhcp(self, ip_address):
        return self.get_attribute_from_ip(ip_address, 'root.Network.Resolver.ObtainFromDHCP') == "yes"  


    def get_firmware_version(self, ip_address):
        return self.get_attribute_from_ip(ip_address, 'root.Properties.Firmware.Version') 

    def get_dhcp_camera(self, network, model_number):
        return Camera.objects.filter(
            network=network,
            model__name=model_number,
            mac_address__isnull=True,
            ip_address__isnull=True
        ).first()                  

    def get_camera(self, mac_address, is_dhcp, ip_address, network):
        try:
            return Camera.objects.get(mac_address=mac_address)
        except ObjectDoesNotExist:    
            if is_dhcp:
                model_number = self.get_attribute_from_ip(ip_address, 'root.Brand.ProdNbr')
                return self.get_dhcp_camera(network, model_number)  
            else:
                return Camera.objects.get(
                    ip_address=ip_address,
                    network=network)   

    def get_xls_row(is_dhcp, ip_address, name):
        if is_dhcp:
            pass
            # get xls_row by name + network  
        else:
            pass
            # get xls_row by ip + network

    def update_xls(self, sheet, ip_address, firmware_version, is_dot1x_disabled, mac_address):        
        for excel_row in sheet.iter_rows(min_row=2, max_col=11, values_only=False):  # Adjust min_row if there's a header
            cell_f = excel_row[5]  # Column F is the 6th column in zero-based index

            if cell_f.value == ip_address:  # Check if the value in F matches the extracted value
                # Assign values to columns K and J (10th and 9th columns)    
                excel_row[8].value = firmware_version   # Column I firmware
                excel_row[9].value = is_dot1x_disabled
                excel_row[10].value = mac_address  # Column K mac address
                break  # Exit after finding the match
        

    def disableHTTPS(self, ip_address):
        self.updateProperty(ip_address, "root.HTTPS.Enable", "no")
        self.updateProperty(ip_address, "root.System.BoaGroupPolicy.admin", "http")
        
    def updateUPnP(self, ip_address, upnp_name):
        self.updateProperty(ip_address, "root.Network.UPnP.FriendlyName", upnp_name)

    def updateAuthMethod(self, ip_address):
         # Turn off 802.1 authentication:
        self.updateProperty(ip_address, "root.Network.Interface.I0.dot1x.Enabled", "no")    

    def configure_device(self, ip_address, upnp_name):
        print("\n\t{}: {}".format(ip_address, upnp_name))
        self.disableHTTPS(ip_address)
        self.updateUPnP(ip_address, upnp_name)
        self.updateAuthMethod(ip_address)   

    def is_dot1x_disabled(self, ip_address):
        # 
        is_dot1x_enabled = self.get_attribute_from_ip(ip_address, 'root.Network.Interface.I0.dot1x.Enabled')
        if is_dot1x_enabled == "no":
            return "Yes"
        elif is_dot1x_enabled == "yes":
            return ""              

    def save_mac_address(self, camera, mac_address):
        camera.mac_address = mac_address 
        camera.save()

    def updateProperty(self, ip_address, property, value):
        text_string = "Set '{}' to '{}'".format(property.replace("root.",""), value)
        url = "http://{}/axis-cgi/param.cgi?action=update&{}={}".format(
            ip_address,
            property,
            value
        )
        response = requests.get(url, auth=HTTPDigestAuth(username, password))
        checkmark = '\u2713'
        if response.status_code == 200:
            print(f"\t\t{GREEN}{checkmark}{RESET} {text_string}")
        else:
            print("Something went wrong...")
            error_pattern = r'<p>(.*?)</p>'
            string_response = str(response._content).replace("\\n", " ").replace("\\", "")
            error_message = re.search(error_pattern, string_response)
            print(error_message.group(1))
            raise SystemExit(0)   

    def load_activeworkbook(self):
        # Load the workbook and select the specific sheet
        try:
            wb = load_workbook(excel_file_path)
        except PermissionError:
            short_file_name = excel_file_path.split("\\")[-1]
            print("\nPermission Error: Please close {} and try again.\n".format(short_file_name))
            raise SystemExit(0)
        return wb

    def handle(self, *args, **options):
        ping_verify = False
        self.clear_screen() 
        bu_identifier = input("Enter business unit ID: \n")             
        
       
        self.print_unassigned_ip_addresses(bu_identifier)
        network = self.get_network_from_bu(bu_identifier)
        wb = self.load_activeworkbook()
        with open(csv_file_path, mode='r', newline='', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile)
            sheet = wb[bu_identifier]
            #skip first row
            next(csv_reader)
            for row in csv_reader:
                # Extract substring from row[2] before the ":"
                ip_address = row[2].split(":")[0]
                mac_address = row[0]
                firmware_version = row[4]
                
               

                # xls_row = self.get_xls_row(is_dhcp, ip_address, name) 
                if ping_verify or self.can_ping(ip_address):
                    ping_verify = True
                    is_dhcp = self.is_dhcp(ip_address)

                    camera = self.get_camera(mac_address, is_dhcp, ip_address, network)   

                    if camera:
                        self.save_mac_address(camera, mac_address)
                        self.configure_device(ip_address, camera.get_upnp_name()) 
                        is_dot1x_disabled = self.is_dot1x_disabled(ip_address)
                        
                        # Iterate through rows in Excel sheet to find the match in column F (6th column, 1-indexed)

                        # if DHCP, search by network + camera name. is_dhcp() to line 171?
                        self.update_xls(
                            sheet, 
                            ip_address, 
                            firmware_version, 
                            is_dot1x_disabled, 
                            mac_address)
                else:
                    print("\nUnable to ping {}".format(ip_address))
                    print("Please ensure new report has been created and network settings are updated.\n")
                    raise SystemExit(0)

        # Save the modified workbook
        wb.save(excel_file_path)
        print("\nExcel file updated successfully.")