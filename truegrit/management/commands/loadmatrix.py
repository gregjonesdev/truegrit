import json
from openpyxl import load_workbook

from django.core.exceptions import ObjectDoesNotExist
from django.core.management.base import BaseCommand
from truegrit.models import (
    Camera,
    CameraManufacturer,
    CameraModel,
    MarketArea,
    Network,
    Project,
    ProjectStatus,
    BusinessUnit,
    StoreChain,
)

class Command(BaseCommand):  

    def create_project(self, number, description):
        current = ProjectStatus.objects.get(name='current')
        try:
            return Project.objects.get(number=number)
        except ObjectDoesNotExist:
            new_project = Project(
                number=number,
                description=description,
                status=current
            )    
            new_project.set_fields_to_base()
            new_project.save()
            return new_project
        
    def create_storechain(self, name):
        try:
            return StoreChain.objects.get(name=name)
        except ObjectDoesNotExist:
            store_chain = StoreChain(
                name=name,
            )    
            store_chain.set_fields_to_base()
            store_chain.save()
            return store_chain
        
    def create_bu(self, identifier, description, market_area):
        try:
            bu = BusinessUnit.objects.get(identifier=identifier)
            bu.market_area = market_area
            bu.save()
            return bu
        except ObjectDoesNotExist:
            bu = BusinessUnit(
                identifier=identifier,
                description=description,
                market_area=market_area,
            )    
            bu.set_fields_to_base()
            bu.save()
            return bu   
         
    def create_marketarea(self, chain, name):
        try:
            return MarketArea.objects.get(
                chain=chain,
                name=name
                )
        except ObjectDoesNotExist:
            print("Create Market Area '{}'".format(name))
            market_area = MarketArea(
                chain=chain,
                name=name
            )    
            market_area.set_fields_to_base()
            market_area.save()
            return market_area   

    def create_axiscameramodel(self, name):
        manufacturer = CameraManufacturer.objects.get(name="Axis")
        try:
            return CameraModel.objects.get(
                name=name, 
                manufacturer=manufacturer)     
        except ObjectDoesNotExist:
            print("Create Axis model: {}".format(name))
            new_model = CameraModel(
                name=name, 
                manufacturer=manufacturer
            )
            new_model.set_fields_to_base()
            new_model.save()
            return new_model
        
    def create_camera(self, network, model, ip_address, name):
        try:
            return Camera.objects.get(
                name=name,
                network=network, 
                ip_address=ip_address,
                model=model)     
        except ObjectDoesNotExist:
            new_camera = Camera(
                name=name, 
                network=network,
                ip_address=ip_address,
                model=model
            )
            new_camera.set_fields_to_base()
            new_camera.save()
            return new_camera    
        
    def createnetwork(self, business_unit, subnet, gateway):
        try:
            return Network.objects.get(
                business_unit=business_unit, 
                subnet=subnet,
                gateway=gateway)     
        except ObjectDoesNotExist:
            new_network = Network(
                business_unit=business_unit, 
                subnet=subnet,
                gateway=gateway
            )
            new_network.set_fields_to_base()
            new_network.save()
            return new_network       
             

    def handle(self, *args, **options):
        # file_path = input("Enter file path of matrix: ")
        file_path = "/Users/gregjones/Downloads/heb-matrix.xlsx"
        print(file_path)
        workbook = load_workbook(filename=file_path)
        project = self.create_project(106191, "HEB 2024 Camera Refresh")
        heb_chain = self.create_storechain("HEB")
        
        #create network 
        # W720 CC - CORPUS CHRISTI TRANSPORTATION
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]  
            cell_value = sheet['A3'].value
            split_cell_value =cell_value.split("-")
            identifier = split_cell_value[0].strip()
            description = split_cell_value[1].strip()
            print(description)
            market_area = self.create_marketarea(heb_chain, sheet['B4'].value.strip())
            business_unit = self.create_bu(identifier, description, market_area)
            subnet = sheet['G4'].value.strip() if sheet['G4'].value else None
            gateway = sheet['H4'].value.strip() if sheet['H4'].value else None
            network = self.createnetwork(business_unit, subnet, gateway)
            for row in sheet.iter_rows(min_row=4, values_only=True):
                # Print data from each column in the current row
                # for cell_value in row:
                #     print(cell_value, end='\t')  # Print cell value with a tab separator
                if row[0]:                             
                    camera_name = row[2].strip()
                    camera_model = self.create_axiscameramodel(row[4].upper().strip()) # camera model
                    print(row[5])
                    ip_address = row[5].strip() if row[5] else None
                    self.create_camera(network, camera_model, ip_address, camera_name)